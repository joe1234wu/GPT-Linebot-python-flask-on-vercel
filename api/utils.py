import os
from typing import List
import uuid
import requests
import tempfile
import logging
import pathlib

from openpyxl import Workbook, load_workbook


TEMPLATE_FILE_PATH = pathlib.Path(__file__).parent.joinpath("template/template.xltx").resolve()
UPLOAD_FILE_URL = "https://tmpfiles.org/api/v1/upload"

def generate_excel_and_upload_wrapper(
    customer_name:str, 
    project_name:str
    ) -> str:
    tempfile_path = os.path.join(tempfile.gettempdir(), str(uuid.uuid4()) + ".xlsx")
    url = ""
    try:
        generate_excel(
            customer_name=customer_name, 
            project_name=project_name, 
            save_file_path=tempfile_path
        )
        url = upload_file_to_cloud(tempfile_path)
    finally:
        logging.info(f"tmpfile_path = {tempfile_path}")
        os.remove(tempfile_path)
    return url

def upload_file_to_cloud(file_path:str) -> str:
    myfiles = {'file': open(file_path ,'rb')}
    r = requests.post(UPLOAD_FILE_URL, files = myfiles)
    if r.status_code != 200:
        raise RuntimeError(r.text)

    r_url = r.json()['data']['url']
    _domain = "tmpfiles.org"
    p1, p2 = r_url.split(_domain)
    return f"{p1}{_domain}/dl{p2}"

def generate_excel(
    customer_name:str, 
    project_name:str, 
    # item_list, 
    save_file_path:str) -> None:

    wb = load_workbook(filename=str(TEMPLATE_FILE_PATH))
    # grab the active worksheet
    ws = wb.active

    ws['A5'] = f"客戶名稱:{customer_name}"
    ws['A6'] = f"工程名稱:{project_name}"

    # Python types will automatically be converted
    import datetime
    now_date = datetime.datetime.now()
    ws['H5'] = f"估價日期: {now_date.year - 1911}.{now_date.month}.{now_date.day}"
    ws['H6'] = f"估價單號: {now_date.strftime('%Y%m%d')}01"

    # Save the file
    wb.template = False
    wb.save(save_file_path)



"""
DATA MODELS
"""
from dataclasses import dataclass, field


@dataclass
class QuoteItem:
    name: str
    quantity: int
    unit: str
    amount: int

@dataclass
class QuoteData: 
    status:int = 1 # 1: wait for cn, 2: wait for pn 
               # 3: during item - name 4: during item - quantity
               # 5: during item - unit 6: during item - amount
               # 7: during item - complete
    customer_name: str = ""
    project_name: str = ""
    items: List[QuoteItem] = field(default_factory=list)

    @property
    def is_ready(self):
        return self.status == 7 
